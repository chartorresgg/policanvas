import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Rutas
ruta_template = r"C:\Users\Charlie\OneDrive - Politécnico Grancolombiano (1)\Repositorio - Planes de Estudio\Template.xlsx"
ruta_destino = r"C:\Users\Charlie\OneDrive - Politécnico Grancolombiano (1)\Repositorio - Planes de Estudio\Planeación de Aulas Master"

# Nombre de la hoja del template que se va a copiar
nombre_hoja_template = "Estructura"  # Puedes ajustar si se llama diferente

# Cargar hoja del template
wb_template = load_workbook(ruta_template)
if nombre_hoja_template not in wb_template.sheetnames:
    print(f"❌ La hoja '{nombre_hoja_template}' no se encontró en el template.")
    exit()

hoja_template = wb_template[nombre_hoja_template]

# Recorremos los archivos destino
procesados = 0
for archivo in os.listdir(ruta_destino):
    if archivo.endswith(".xlsx") and not archivo.startswith("~$"):
        ruta_archivo = os.path.join(ruta_destino, archivo)
        try:
            wb = load_workbook(ruta_archivo)
            # Si ya existe una hoja con ese nombre, la eliminamos
            if nombre_hoja_template in wb.sheetnames:
                std = wb[nombre_hoja_template]
                wb.remove(std)

            # Crear nueva hoja y copiar contenido del template
            nueva_hoja = wb.create_sheet(title=nombre_hoja_template)

            for i, fila in enumerate(hoja_template.iter_rows(values_only=True), start=1):
                for j, valor in enumerate(fila, start=1):
                    celda = f"{get_column_letter(j)}{i}"
                    nueva_hoja[celda] = valor

            # Guardar el archivo modificado
            wb.save(ruta_archivo)
            procesados += 1
            print(f"✅ Añadida hoja '{nombre_hoja_template}' a: {archivo}")
        except Exception as e:
            print(f"❌ Error procesando {archivo}: {e}")

print(f"\n🎉 Hoja '{nombre_hoja_template}' copiada a {procesados} archivos en total.")
